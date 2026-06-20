#!/usr/bin/env python3
"""Extract the provided SwissHacks xlsx dataset into JSON seeds the backend loads.

Source files live in SIX-Noumena-NTT-Data/data/. This is the *provided dataset*
(not a live API), so parsing it locally is the intended path. Re-run whenever the
xlsx files change:  python3 scripts/extract-data.py
"""
import json
import os
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, "SIX-Noumena-NTT-Data", "data")
OUT = os.path.join(ROOT, "backend", "data", "seeds")
os.makedirs(os.path.join(OUT, "crm"), exist_ok=True)
os.makedirs(os.path.join(OUT, "portfolios"), exist_ok=True)


def iso(v):
    return v.isoformat() if hasattr(v, "isoformat") else (str(v) if v is not None else None)


def rows(ws, headers):
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(c is None for c in r):
            continue
        out.append({h: iso(r[i]) if i < len(r) else None for i, h in enumerate(headers)})
    return out


# --- CRM (one tab per client) -------------------------------------------------
crm_wb = openpyxl.load_workbook(os.path.join(DATA, "SwissHacks CRM.xlsx"), data_only=True)
CRM_KEYS = {"CRM Raeber": "raeber", "CRM Schneider": "schneider",
            "CRM Huber": "huber", "CRM Ammann": "ammann"}
for sheet, cid in CRM_KEYS.items():
    ws = crm_wb[sheet]
    notes = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[4]:
            continue
        notes.append({"date": iso(r[0]), "medium": r[1], "rm": r[2],
                      "contact": r[3], "note": r[4]})
    with open(os.path.join(OUT, "crm", f"{cid}.json"), "w") as f:
        json.dump(notes, f, ensure_ascii=False, indent=2)
    print(f"crm/{cid}.json  ({len(notes)} notes)")

# --- Portfolios + CIO list ----------------------------------------------------
pf_wb = openpyxl.load_workbook(os.path.join(DATA, "SwissHacks Portfolio Construction.xlsx"), data_only=True)

PF_HEAD = ["assetClass", "subAssetClass", "region", "industry", "issuer",
           "security", "isin", "targetChf", "currentChf", "valor", "mic", "yahoo"]
for sheet, mandate in {"Sample Portfolio Defensive": "defensive",
                       "Sample Portfolio Balanced": "balanced",
                       "Sample Portfolio Growth": "growth"}.items():
    data = rows(pf_wb[sheet], PF_HEAD)
    with open(os.path.join(OUT, "portfolios", f"{mandate}.json"), "w") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"portfolios/{mandate}.json  ({len(data)} positions)")

CIO_HEAD = ["rating", "ratingSince", "assetClass", "subAssetClass", "region",
            "industry", "issuer", "security", "isin", "cioView", "valor", "mic", "yahoo", "asOf"]
cio = rows(pf_wb["CIO Recommendation List"], CIO_HEAD)
with open(os.path.join(OUT, "cio.json"), "w") as f:
    json.dump(cio, f, ensure_ascii=False, indent=2)
print(f"cio.json  ({len(cio)} recommendations)")

print("\nDone.")
